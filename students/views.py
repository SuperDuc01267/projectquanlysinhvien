from datetime import datetime, timedelta
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.views import LoginView, LogoutView
from django.core.mail import EmailMessage
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views.generic import CreateView, DeleteView, ListView, TemplateView, UpdateView
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .forms import (
    ExamScheduleForm,
    ScheduleForm,
    ScoreForm,
    StudentAuthenticationForm,
    StudentForm,
    StudentProfileForm,
    SubjectForm,
)
from .models import ExamSchedule, Schedule, Score, Student, Subject


TIMETABLE_DAYS = [
    ("2", "Thứ 2"),
    ("3", "Thứ 3"),
    ("4", "Thứ 4"),
    ("5", "Thứ 5"),
    ("6", "Thứ 6"),
    ("7", "Thứ 7"),
    ("CN", "Chủ nhật"),
]

TIMETABLE_PERIODS = [
    {"number": 1, "label": "Tiết 1", "start": "07:00", "end": "07:50"},
    {"number": 2, "label": "Tiết 2", "start": "07:50", "end": "08:40"},
    {"number": 3, "label": "Tiết 3", "start": "08:50", "end": "09:40"},
    {"number": 4, "label": "Tiết 4", "start": "09:40", "end": "10:30"},
    {"number": 5, "label": "Tiết 5", "start": "10:40", "end": "11:30"},
    {"number": 6, "label": "Tiết 6", "start": "13:00", "end": "13:50"},
    {"number": 7, "label": "Tiết 7", "start": "13:50", "end": "14:40"},
    {"number": 8, "label": "Tiết 8", "start": "14:50", "end": "15:40"},
    {"number": 9, "label": "Tiết 9", "start": "15:40", "end": "16:30"},
    {"number": 10, "label": "Tiết 10", "start": "16:40", "end": "17:30"},
]


def _time_to_minutes(value):
    return value.hour * 60 + value.minute


def _slot_to_minutes(slot_value):
    hour, minute = map(int, slot_value.split(":"))
    return hour * 60 + minute


def build_timetable_cards(schedules):
    day_columns = {code: index + 2 for index, (code, _) in enumerate(TIMETABLE_DAYS)}
    period_bounds = [
        {
            "number": period["number"],
            "start": _slot_to_minutes(period["start"]),
            "end": _slot_to_minutes(period["end"]),
        }
        for period in TIMETABLE_PERIODS
    ]
    cards = []
    for item in schedules:
        start_minutes = _time_to_minutes(item.start_time)
        end_minutes = _time_to_minutes(item.end_time)
        overlapping = [
            period["number"]
            for period in period_bounds
            if start_minutes < period["end"] and end_minutes > period["start"]
        ]
        if overlapping:
            start_period = min(overlapping)
            end_period = max(overlapping)
        else:
            start_period = 1
            end_period = 1
        cards.append(
            {
                "item": item,
                "grid_column": day_columns.get(item.weekday, 2),
                "grid_row": start_period + 1,
                "row_span": max(1, end_period - start_period + 1),
            }
        )
    return cards


def get_selected_date(request):
    raw_value = request.GET.get("date")
    if raw_value:
        try:
            return datetime.strptime(raw_value, "%Y-%m-%d").date()
        except ValueError:
            return timezone.localdate()
    return timezone.localdate()


def build_excel_response(filename, title, metadata_rows, headers, rows):
    content = build_excel_bytes(title, metadata_rows, headers, rows)
    response = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_excel_bytes(title, metadata_rows, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Du lieu"

    title_fill = PatternFill("solid", fgColor="0F766E")
    header_fill = PatternFill("solid", fgColor="D9EDE8")
    white_font = Font(color="FFFFFF", bold=True, size=14)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin_side = Side(style="thin", color="B8C4D4")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    last_col = chr(ord("A") + len(headers) - 1)
    sheet.merge_cells(f"A1:{last_col}1")
    sheet["A1"] = title
    sheet["A1"].fill = title_fill
    sheet["A1"].font = white_font
    sheet["A1"].alignment = center
    sheet.row_dimensions[1].height = 26

    current_row = 3
    for label, value in metadata_rows:
        sheet[f"A{current_row}"] = label
        sheet[f"A{current_row}"].font = bold_font
        sheet[f"B{current_row}"] = value
        current_row += 1

    current_row += 1
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=current_row, column=column_index, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for data_row in rows:
        current_row += 1
        for column_index, value in enumerate(data_row, start=1):
            cell = sheet.cell(row=current_row, column=column_index, value=value)
            cell.border = border
            cell.alignment = center if column_index != 2 else left

    for column_index in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(column_index)
        for cell in sheet[column_letter]:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except Exception:
                pass
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 36)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


def build_pdf_bytes(title, metadata_rows, headers, rows):
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    for label, value in metadata_rows:
        story.append(Paragraph(f"<b>{label}:</b> {value}", styles["Normal"]))
    story.append(Spacer(1, 12))

    table_data = [headers] + rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EDE8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#92A4B8")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ]
        )
    )
    story.append(table)
    document.build(story)
    output.seek(0)
    return output.read()


def send_export_email(student, mail_subject, filename, content, mimetype):
    if not student.email:
        return False, "Sinh vien chua co email de nhan file."
    try:
        message = EmailMessage(
            subject=mail_subject,
            body=(
                f"Xin chao {student.full_name},\n\n"
                f"He thong vua tao file '{filename}' cho ban. File duoc dinh kem trong email nay.\n\n"
                "Truong hop ban khong yeu cau thao tac nay, vui long lien he quan tri vien."
            ),
            from_email="Student Management System <nguyenduc01267@gmail.com>",
            to=[student.email],
        )
        message.attach(filename, content, mimetype)
        message.send(fail_silently=False)
        return True, ""
    except Exception as exc:
        return False, str(exc)


class HomeView(TemplateView):
    template_name = "students/home.html"

    def dispatch(self, request, *args, **kwargs):
        # Chỉ redirect admin khi vào /admin/, còn vào trang chủ thì luôn hiện trang chủ cho mọi user
        return super().dispatch(request, *args, **kwargs)


class SuccessMessageMixin:
    success_message = ""

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.success_message:
            messages.success(self.request, self.success_message)
        return response


class StaffRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        if self.request.user.is_authenticated and not self.request.user.is_staff:
            messages.error(self.request, "Bạn không có quyền truy cập khu quản trị.")
            return redirect("students:student_result")
        if self.request.user.is_authenticated and self.request.user.is_staff:
            return redirect("/admin/")
        return super().handle_no_permission()


class StudentPortalRequiredMixin(LoginRequiredMixin):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if request.user.is_staff:
            return redirect("/admin/")
        self.student = Student.objects.filter(user=request.user).first()
        if self.student is None:
            logout(request)
            messages.error(
                request,
                "Tài khoản này chưa được liên kết với hồ sơ sinh viên. Vui lòng liên hệ quản trị viên.",
            )
            return redirect("students:login")
        return super().dispatch(request, *args, **kwargs)


class StudentLoginView(LoginView):
    template_name = "students/login.html"
    authentication_form = StudentAuthenticationForm
    redirect_authenticated_user = True

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if request.user.is_staff:
                return redirect("/admin/")
            if not Student.objects.filter(user=request.user).exists():
                logout(request)
                messages.error(
                    request,
                    "Tài khoản này chưa được liên kết với hồ sơ sinh viên. Vui lòng liên hệ quản trị viên.",
                )
                return redirect("students:login")
            return redirect("students:student_result")
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.user.is_staff:
            return redirect("/admin/")
        if not Student.objects.filter(user=self.request.user).exists():
            logout(self.request)
            messages.error(
                self.request,
                "Tài khoản này chưa được liên kết với hồ sơ sinh viên. Vui lòng liên hệ quản trị viên.",
            )
            return redirect("students:login")
        return response

    def get_success_url(self):
        return reverse_lazy("students:student_result")


class StudentLogoutView(LogoutView):
    def get_next_page(self):
        return reverse_lazy("students:home")

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"
        return response


def admin_portal_logout(request):
    logout(request)
    response = redirect("/admin/login/")
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


class StudentListView(StaffRequiredMixin, ListView):
    model = Student
    template_name = "students/student_list.html"
    context_object_name = "students"


class StudentCreateView(StaffRequiredMixin, SuccessMessageMixin, CreateView):
    model = Student
    form_class = StudentForm
    template_name = "students/student_form.html"
    success_url = reverse_lazy("students:student_list")
    success_message = "Thêm sinh viên thành công."


class StudentUpdateView(StaffRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = "students/student_form.html"
    success_url = reverse_lazy("students:student_list")
    success_message = "Cập nhật sinh viên thành công."


class StudentDeleteView(StaffRequiredMixin, DeleteView):
    model = Student
    template_name = "students/confirm_delete.html"
    success_url = reverse_lazy("students:student_list")

    def form_valid(self, form):
        messages.success(self.request, "Xóa sinh viên thành công.")
        return super().form_valid(form)


class SubjectListView(StaffRequiredMixin, ListView):
    model = Subject
    template_name = "students/subject_list.html"
    context_object_name = "subjects"


class SubjectCreateView(StaffRequiredMixin, SuccessMessageMixin, CreateView):
    model = Subject
    form_class = SubjectForm
    template_name = "students/subject_form.html"
    success_url = reverse_lazy("students:subject_list")
    success_message = "Thêm môn học thành công."


class SubjectUpdateView(StaffRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Subject
    form_class = SubjectForm
    template_name = "students/subject_form.html"
    success_url = reverse_lazy("students:subject_list")
    success_message = "Cập nhật môn học thành công."


class SubjectDeleteView(StaffRequiredMixin, DeleteView):
    model = Subject
    template_name = "students/confirm_delete.html"
    success_url = reverse_lazy("students:subject_list")

    def form_valid(self, form):
        messages.success(self.request, "Xóa môn học thành công.")
        return super().form_valid(form)


class ScoreListView(StaffRequiredMixin, ListView):
    model = Score
    template_name = "students/score_list.html"
    context_object_name = "scores"

    def get_queryset(self):
        return (
            Score.objects.select_related("student", "subject")
            .order_by("student__student_code", "subject__subject_code")
        )


class ScoreCreateView(StaffRequiredMixin, SuccessMessageMixin, CreateView):
    model = Score
    form_class = ScoreForm
    template_name = "students/score_form.html"
    success_url = reverse_lazy("students:score_list")
    success_message = "Thêm bảng điểm thành công."


class ScoreUpdateView(StaffRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Score
    form_class = ScoreForm
    template_name = "students/score_form.html"
    success_url = reverse_lazy("students:score_list")
    success_message = "Cập nhật bảng điểm thành công."


class ScoreDeleteView(StaffRequiredMixin, DeleteView):
    model = Score
    template_name = "students/confirm_delete.html"
    success_url = reverse_lazy("students:score_list")

    def form_valid(self, form):
        messages.success(self.request, "Xóa bảng điểm thành công.")
        return super().form_valid(form)


class ScheduleListView(StaffRequiredMixin, ListView):
    model = Schedule
    template_name = "students/schedule_list.html"
    context_object_name = "schedules"

    def get_queryset(self):
        return (
            Schedule.objects.select_related("student", "subject")
            .order_by("student__student_code", "weekday", "start_time")
        )


class ScheduleCreateView(StaffRequiredMixin, SuccessMessageMixin, CreateView):
    model = Schedule
    form_class = ScheduleForm
    template_name = "students/schedule_form.html"
    success_url = reverse_lazy("students:schedule_list")
    success_message = "Thêm thời khóa biểu thành công."


class ScheduleUpdateView(StaffRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Schedule
    form_class = ScheduleForm
    template_name = "students/schedule_form.html"
    success_url = reverse_lazy("students:schedule_list")
    success_message = "Cập nhật thời khóa biểu thành công."


class ScheduleDeleteView(StaffRequiredMixin, DeleteView):
    model = Schedule
    template_name = "students/confirm_delete.html"
    success_url = reverse_lazy("students:schedule_list")

    def form_valid(self, form):
        messages.success(self.request, "Xóa thời khóa biểu thành công.")
        return super().form_valid(form)


class ExamScheduleListView(StaffRequiredMixin, ListView):
    model = ExamSchedule
    template_name = "students/exam_schedule_list.html"
    context_object_name = "exam_schedules"

    def get_queryset(self):
        return (
            ExamSchedule.objects.select_related("student", "subject")
            .order_by("student__student_code", "exam_date", "start_time")
        )


class ExamScheduleCreateView(StaffRequiredMixin, SuccessMessageMixin, CreateView):
    model = ExamSchedule
    form_class = ExamScheduleForm
    template_name = "students/exam_schedule_form.html"
    success_url = reverse_lazy("students:exam_schedule_list")
    success_message = "Thêm lịch thi thành công."


class ExamScheduleUpdateView(StaffRequiredMixin, SuccessMessageMixin, UpdateView):
    model = ExamSchedule
    form_class = ExamScheduleForm
    template_name = "students/exam_schedule_form.html"
    success_url = reverse_lazy("students:exam_schedule_list")
    success_message = "Cập nhật lịch thi thành công."


class ExamScheduleDeleteView(StaffRequiredMixin, DeleteView):
    model = ExamSchedule
    template_name = "students/confirm_delete.html"
    success_url = reverse_lazy("students:exam_schedule_list")

    def form_valid(self, form):
        messages.success(self.request, "Xóa lịch thi thành công.")
        return super().form_valid(form)


class StudentResultView(StudentPortalRequiredMixin, TemplateView):
    template_name = "students/student_result.html"

    def get_scores(self):
        return (
            Score.objects.select_related("subject")
            .filter(student=self.student)
            .order_by("subject__subject_code")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        scores = self.get_scores()
        context.update(
            {
                "student": self.student,
                "scores": scores,
                "average_score": self.student.average_score(),
                "classification": self.student.academic_classification(),
            }
        )
        return context


class StudentProfileView(StudentPortalRequiredMixin, UpdateView):
    model = Student
    form_class = StudentProfileForm
    template_name = "students/student_profile.html"
    success_url = reverse_lazy("students:student_profile")

    def get_object(self, queryset=None):
        return self.student

    def form_valid(self, form):
        messages.success(self.request, "Cap nhat thong tin ca nhan thanh cong.")
        return super().form_valid(form)


class StudentResultExcelView(StudentPortalRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        scores = (
            Score.objects.select_related("subject")
            .filter(student=self.student)
            .order_by("subject__subject_code")
        )
        metadata_rows = [
            ("Ho ten", self.student.full_name),
            ("Ma sinh vien", self.student.student_code),
            ("Lop", self.student.class_name),
            ("Email", self.student.email),
            ("Diem trung binh", self.student.average_score()),
            ("Hoc luc", self.student.academic_classification()),
            ("Ngay xuat", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
        ]
        headers = ["Ma mon", "Mon hoc", "Qua trinh", "Giua ky", "Cuoi ky", "Tong ket"]
        rows = [
            [
                score.subject.subject_code,
                score.subject.name,
                float(score.attendance_score),
                float(score.midterm_score),
                float(score.final_exam_score),
                float(score.final_score),
            ]
            for score in scores
        ]
        filename = f"ket-qua-hoc-tap-{self.student.student_code}.xlsx"
        content = build_excel_bytes("PHIEU KET QUA HOC TAP", metadata_rows, headers, rows)
        email_sent, email_error = send_export_email(
            self.student,
            "Ket qua hoc tap dang dinh kem",
            filename,
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if email_sent:
            messages.success(request, f"Da gui file Excel vao email {self.student.email}.")
        else:
            messages.error(request, f"Khong gui duoc email: {email_error}")
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class StudentResultPdfView(StudentPortalRequiredMixin, TemplateView):
    template_name = "students/student_result_pdf.html"

    def get(self, request, *args, **kwargs):
        scores = (
            Score.objects.select_related("subject")
            .filter(student=self.student)
            .order_by("subject__subject_code")
        )
        metadata_rows = [
            ("Ho ten", self.student.full_name),
            ("Ma sinh vien", self.student.student_code),
            ("Lop", self.student.class_name),
            ("Email", self.student.email),
            ("Diem trung binh", self.student.average_score()),
            ("Hoc luc", self.student.academic_classification()),
            ("Ngay xuat", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
        ]
        headers = ["Ma mon", "Mon hoc", "Qua trinh", "Giua ky", "Cuoi ky", "Tong ket"]
        rows = [
            [
                score.subject.subject_code,
                score.subject.name,
                str(score.attendance_score),
                str(score.midterm_score),
                str(score.final_exam_score),
                str(score.final_score),
            ]
            for score in scores
        ]
        pdf_bytes = build_pdf_bytes("PHIEU KET QUA HOC TAP", metadata_rows, headers, rows)
        email_sent, email_error = send_export_email(
            self.student,
            "Phieu ket qua hoc tap dang dinh kem",
            f"ket-qua-hoc-tap-{self.student.student_code}.pdf",
            pdf_bytes,
            "application/pdf",
        )
        if email_sent:
            messages.success(request, f"Da gui file PDF vao email {self.student.email}.")
        else:
            messages.error(request, f"Khong gui duoc email: {email_error}")
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        scores = (
            Score.objects.select_related("subject")
            .filter(student=self.student)
            .order_by("subject__subject_code")
        )
        context.update(
            {
                "student": self.student,
                "scores": scores,
                "average_score": self.student.average_score(),
                "classification": self.student.academic_classification(),
                "generated_at": timezone.localtime(),
            }
        )
        return context


class StudentScheduleView(StudentPortalRequiredMixin, TemplateView):
    template_name = "students/student_schedule.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_date = get_selected_date(self.request)
        week_start = selected_date - timedelta(days=selected_date.weekday())
        week_dates = []
        for offset, (day_code, day_label) in enumerate(TIMETABLE_DAYS):
            current_date = week_start + timedelta(days=offset)
            week_dates.append(
                {
                    "code": day_code,
                    "label": day_label,
                    "date": current_date,
                    "is_selected": current_date == selected_date,
                }
            )
        schedules = (
            Schedule.objects.select_related("subject")
            .filter(student=self.student)
            .filter(effective_from__lte=selected_date, effective_to__gte=selected_date)
            .order_by("weekday", "start_time")
        )
        context.update(
            {
                "student": self.student,
                "schedules": schedules,
                "timetable_days": TIMETABLE_DAYS,
                "timetable_periods": TIMETABLE_PERIODS,
                "timetable_cards": build_timetable_cards(schedules),
                "selected_date": selected_date,
                "week_dates": week_dates,
                "week_start": week_dates[0]["date"],
                "week_end": week_dates[-1]["date"],
            }
        )
        return context


class StudentScheduleExcelView(StudentPortalRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        selected_date = get_selected_date(request)
        schedules = (
            Schedule.objects.select_related("subject")
            .filter(student=self.student)
            .filter(effective_from__lte=selected_date, effective_to__gte=selected_date)
            .order_by("weekday", "start_time")
        )
        metadata_rows = [
            ("Ho ten", self.student.full_name),
            ("Ma sinh vien", self.student.student_code),
            ("Ngay xem lich", selected_date.strftime("%d/%m/%Y")),
            ("Ngay xuat", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
        ]
        headers = ["Thu", "Mon hoc", "Phong", "Bat dau", "Ket thuc", "Hieu luc tu", "Hieu luc den", "Ghi chu"]
        rows = [
            [
                item.get_weekday_display(),
                item.subject.name,
                item.room,
                item.start_time.strftime("%H:%M"),
                item.end_time.strftime("%H:%M"),
                item.effective_from.strftime("%d/%m/%Y"),
                item.effective_to.strftime("%d/%m/%Y"),
                item.note,
            ]
            for item in schedules
        ]
        filename = f"thoi-khoa-bieu-{self.student.student_code}.xlsx"
        content = build_excel_bytes("THOI KHOA BIEU", metadata_rows, headers, rows)
        email_sent, email_error = send_export_email(
            self.student,
            "Thoi khoa bieu dang dinh kem",
            filename,
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if email_sent:
            messages.success(request, f"Da gui file Excel vao email {self.student.email}.")
        else:
            messages.error(request, f"Khong gui duoc email: {email_error}")
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class StudentSchedulePdfView(StudentPortalRequiredMixin, TemplateView):
    template_name = "students/student_schedule_pdf.html"

    def get(self, request, *args, **kwargs):
        selected_date = get_selected_date(request)
        schedules = (
            Schedule.objects.select_related("subject")
            .filter(student=self.student)
            .filter(effective_from__lte=selected_date, effective_to__gte=selected_date)
            .order_by("weekday", "start_time")
        )
        metadata_rows = [
            ("Ho ten", self.student.full_name),
            ("Ma sinh vien", self.student.student_code),
            ("Ngay xem lich", selected_date.strftime("%d/%m/%Y")),
            ("Ngay xuat", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
        ]
        headers = ["Thu", "Mon hoc", "Phong", "Bat dau", "Ket thuc", "Hieu luc"]
        rows = [
            [
                item.get_weekday_display(),
                item.subject.name,
                item.room,
                item.start_time.strftime("%H:%M"),
                item.end_time.strftime("%H:%M"),
                f"{item.effective_from.strftime('%d/%m/%Y')} - {item.effective_to.strftime('%d/%m/%Y')}",
            ]
            for item in schedules
        ]
        pdf_bytes = build_pdf_bytes("THOI KHOA BIEU", metadata_rows, headers, rows)
        email_sent, email_error = send_export_email(
            self.student,
            "Thoi khoa bieu dang dinh kem",
            f"thoi-khoa-bieu-{self.student.student_code}.pdf",
            pdf_bytes,
            "application/pdf",
        )
        if email_sent:
            messages.success(request, f"Da gui file PDF vao email {self.student.email}.")
        else:
            messages.error(request, f"Khong gui duoc email: {email_error}")
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_date = get_selected_date(self.request)
        schedules = (
            Schedule.objects.select_related("subject")
            .filter(student=self.student)
            .filter(effective_from__lte=selected_date, effective_to__gte=selected_date)
            .order_by("weekday", "start_time")
        )
        context.update(
            {
                "student": self.student,
                "selected_date": selected_date,
                "schedules": schedules,
                "generated_at": timezone.localtime(),
            }
        )
        return context


class StudentExamScheduleView(StudentPortalRequiredMixin, TemplateView):
    template_name = "students/student_exam_schedule.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        exam_schedules = (
            ExamSchedule.objects.select_related("subject")
            .filter(student=self.student)
            .order_by("exam_date", "start_time")
        )
        context.update({"student": self.student, "exam_schedules": exam_schedules})
        return context


class StudentExamScheduleExcelView(StudentPortalRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        exam_schedules = (
            ExamSchedule.objects.select_related("subject")
            .filter(student=self.student)
            .order_by("exam_date", "start_time")
        )
        metadata_rows = [
            ("Ho ten", self.student.full_name),
            ("Ma sinh vien", self.student.student_code),
            ("Ngay xuat", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
        ]
        headers = ["Ngay thi", "Gio thi", "Mon hoc", "Phong thi", "So bao danh", "Ghi chu"]
        rows = [
            [
                item.exam_date.strftime("%d/%m/%Y"),
                item.start_time.strftime("%H:%M"),
                item.subject.name,
                item.room,
                item.seat_number,
                item.note,
            ]
            for item in exam_schedules
        ]
        filename = f"lich-thi-{self.student.student_code}.xlsx"
        content = build_excel_bytes("LICH THI", metadata_rows, headers, rows)
        email_sent, email_error = send_export_email(
            self.student,
            "Lich thi dang dinh kem",
            filename,
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if email_sent:
            messages.success(request, f"Da gui file Excel vao email {self.student.email}.")
        else:
            messages.error(request, f"Khong gui duoc email: {email_error}")
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class StudentExamSchedulePdfView(StudentPortalRequiredMixin, TemplateView):
    template_name = "students/student_exam_schedule_pdf.html"

    def get(self, request, *args, **kwargs):
        exam_schedules = (
            ExamSchedule.objects.select_related("subject")
            .filter(student=self.student)
            .order_by("exam_date", "start_time")
        )
        metadata_rows = [
            ("Ho ten", self.student.full_name),
            ("Ma sinh vien", self.student.student_code),
            ("Ngay xuat", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
        ]
        headers = ["Ngay thi", "Gio thi", "Mon hoc", "Phong thi", "So bao danh", "Ghi chu"]
        rows = [
            [
                item.exam_date.strftime("%d/%m/%Y"),
                item.start_time.strftime("%H:%M"),
                item.subject.name,
                item.room,
                item.seat_number,
                item.note or "-",
            ]
            for item in exam_schedules
        ]
        pdf_bytes = build_pdf_bytes("LICH THI", metadata_rows, headers, rows)
        email_sent, email_error = send_export_email(
            self.student,
            "Lich thi dang dinh kem",
            f"lich-thi-{self.student.student_code}.pdf",
            pdf_bytes,
            "application/pdf",
        )
        if email_sent:
            messages.success(request, f"Da gui file PDF vao email {self.student.email}.")
        else:
            messages.error(request, f"Khong gui duoc email: {email_error}")
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        exam_schedules = (
            ExamSchedule.objects.select_related("subject")
            .filter(student=self.student)
            .order_by("exam_date", "start_time")
        )
        context.update(
            {
                "student": self.student,
                "exam_schedules": exam_schedules,
                "generated_at": timezone.localtime(),
            }
        )
        return context
